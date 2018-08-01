from django.shortcuts import render, redirect
from .models import *
from .forms import *
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.contrib.auth.views import logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie,csrf_exempt
from django.views.decorators.csrf import requires_csrf_token
# Create your views here.
from django.views.defaults import page_not_found

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from openpyxl.drawing.image import Image
# Libreria para acceder a la fecha actual
from datetime import *



@login_required(None,'login','/login/')
def index(request):
    ultimo_contrato = Contrato.objects.latest('id')
    tracks = Track.objects.all()
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        # obtener todos los repositores que marcaron presencia hoy
        hoy = date.today()
        # Intentamos obtener cuantos usuarios hicieron sus marcaciones en el dia.
        try:
            marcantes_entrada = Marcacione.objects.filter(fecha=hoy).distinct('usuario').count()
        except:
            marcantes_entrada = 0
        # marcantes_salida = Marcacione.objects.filter(fecha=hoy,estado="1").distinct('usuario').count()
        marcaciones= Marcacione.objects.all().order_by('-id')[:10]
        marcaciones_todas= Marcacione.objects.all().count()
        usuarios = User.objects.filter(is_active=True,is_superuser=False, is_staff=False)
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(usuarios.count()*100/limite_usuarios)
        return render(request, 'index.html',{'marcaciones':marcaciones,'marcaciones_todas':marcaciones_todas,'usuarios':usuarios,
                                                'porcentaje':porcentaje,'limite_usuarios':limite_usuarios, 'marcantes_entrada':marcantes_entrada,
                                                'tracks':tracks })
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})

#@csrf_protect
#@ensure_csrf_cookie	
def login_view(request):
    state = ""
    if request.method == "POST":
        print("HIZO EL POST")
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_staff:
                print(user)
                ultimo_contrato = Contrato.objects.latest('id')
                hoy = date.today()
                login(request, user)
                state = "Te has logueado correctamente!"
                if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
                    return redirect('/')
                else:
                    logout(request)
                    state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
                    print(state)
                    return render(request,'login.html',{'state':state})
            else:
                state = "Tu cuenta no esta activa, por favor ponte en contacto con el administrador."
        else:
            state = "Tu nombre de usuario y/o contraseña no coinciden."
    if request.user.is_authenticated():
        ultimo_contrato = Contrato.objects.latest('id')
        hoy = date.today()
        if ultimo_contrato.fecha_caducidad >= hoy:
            return redirect('/')
        else:
            print('77')
            logout(request)
            state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
            return render(request,'login.html',{'state':state})		
    return render(request,'login.html',{'state':state})

@login_required(None, 'login', '/login/')
#@csrf_protect
#@ensure_csrf_cookie
def logout_view(request):
    logout(request)
    return redirect("/")
	
@login_required(None,'login','/login/')	
def registrar_repositor(request):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        usuarios = User.objects.filter(is_active=True,is_superuser=False, is_staff=False)
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(usuarios.count()*100/limite_usuarios)
        if ultimo_contrato.limite_usuarios >= usuarios.count():
            if request.method == "POST":
                form = RegistroRepositorForm(request.POST)
                if form.is_valid():
                    new_user = User.objects.create_user(form.cleaned_data['username'], (form.cleaned_data['email'] or "controller@tecnodesign.com.py"), form.cleaned_data['username'])
                    print(new_user.id)
                    new_user.first_name = form.cleaned_data['nombres']
                    new_user.last_name = form.cleaned_data['apellidos']
                    new_user.save()
                    #jefe_supermercados = request.POST.get('jefe_supermercado')
                    #Obtenemos el id del usuario para crear su perfil
                    usuario1= User.objects.latest('id')
                    nuevo_personal = Usuario.objects.create(usuario=usuario1,legajo=form.cleaned_data['legajo'])
                    return redirect('/listado_repositores')
                else:
                    return render(request,'registrar_admin.html',{'form':form,'usuarios':usuarios,'porcentaje':porcentaje})
            else:
                form = RegistroRepositorForm()
                
            return render(request, 'registrar_repositor.html',{'form':form,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
        else:
            return redirect('/')
            
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})

@login_required(None,'login','/login/')
def add_comercio(request):
    if request.method == "POST":
        nombre_comercio =   request.POST.get('nombre')
        coordenadas =   request.POST.get('array')
        print(coordenadas)
        nuevo_personal = Comercio.objects.create(nombre=nombre_comercio,array=coordenadas)
        return redirect('/comercios/listado/')
    return render(request,'comercios/add.html')

@login_required(None,'login','/login/')
def listado_comercio(request):
    comercios = Comercio.objects.all()
    return render(request,'comercios/listado.html',{'comercios':comercios})


@login_required(None,'login','/login/')	
def registrar_admin(request):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        usuarios = User.objects.filter(is_active=True,is_superuser=False, is_staff=False)
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(usuarios.count()*100/limite_usuarios)
        if ultimo_contrato.limite_usuarios >= usuarios.count():
            if request.method == "POST":
                form = RegistroRepositorForm(request.POST)
                if form.is_valid():
                    new_user = User.objects.create_user(form.cleaned_data['username'], (form.cleaned_data['email'] or "controller@tecnodesign.com.py"), form.cleaned_data['username'])
                    print(new_user.id)
                    new_user.first_name = form.cleaned_data['nombres']
                    new_user.last_name = form.cleaned_data['apellidos']
                    new_user.is_staff = True
                    new_user.save()
                    #jefe_supermercados = request.POST.get('jefe_supermercado')
                    #Obtenemos el id del usuario para crear su perfil
                    usuario1= User.objects.latest('id')
                    nuevo_personal = Usuario.objects.create(usuario=usuario1,legajo=form.cleaned_data['legajo'])
                    return redirect('/listado_repositores')
                else:
                    return render(request,'registrar_admin.html',{'form':form,'usuarios':usuarios,'porcentaje':porcentaje})
            else:
                form = RegistroRepositorForm()
                
            return render(request, 'registrar_admin.html',{'form':form,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
        else:
            return redirect('/')
            
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})


@login_required(None,'login','/login/')
def listado_repositores(request):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        repositores = Usuario.objects.all()
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(repositores.count()*100/limite_usuarios)
        return render(request, 'listado_repositores.html',{'repositores':repositores,'porcentaje':porcentaje})
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        return render(request,'login.html',{'state':state})

@login_required(None,'login','/login/')
def activos(request):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        repositores = Usuario.objects.all()
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(repositores.count()*100/limite_usuarios)
        return render(request, 'activos.html',{'usuarios':repositores,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        return render(request,'login.html',{'state':state})

@login_required(None,'login','/login/')
def inactivos(request):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        repositores = Usuario.objects.all()
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(repositores.count()*100/limite_usuarios)
        return render(request, 'inactivos.html',{'usuarios':repositores,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        return render(request,'login.html',{'state':state})

@login_required(None,'login','/login/') 
def editar_comercio(request,comercio):
    if request.method == "POST":
        nombre_comercio =   request.POST.get('nombre')
        coordenadas =   request.POST.get('array')
        print("Coordenadas : "+str(coordenadas))
        if(coordenadas != ''):
            nuevo_personal = Comercio.objects.filter(id=comercio).update(nombre=nombre_comercio,array=coordenadas)
        else:
            nuevo_personal = Comercio.objects.filter(id=comercio).update(nombre=nombre_comercio)
        return redirect('/comercios/listado/')
    else:
        user_repositor = Comercio.objects.get(id=comercio)
        form = EditarComercioForm(instance=user_repositor)

    comercio = Comercio.objects.get(id=comercio)
    return render(request, 'comercios/editar.html',{'form':form,'comercio':comercio})

@login_required(None,'login','/login/') 
def eliminar_comercio(request,comercio):
    print(comercio)
    if comercio != '':
        Comercio.objects.get(id=comercio).delete()
    return redirect('/comercios/listado/')

@login_required(None,'login','/login/')	
def editar_repositor(request,usuario):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        usuarios = User.objects.filter(is_active=True,is_superuser=False, is_staff=False)
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(usuarios.count()*100/limite_usuarios)
        if request.method == "POST":
            jefe_supermercados = request.POST.get('jefe_supermercado')
            nombres = request.POST.get('nombres')
            apellidos = request.POST.get('apellidos')
            email = request.POST.get('email')
            legajo = request.POST.get('legajo')
            usuario1 = request.POST.get('usuario')
            exclude_id = Usuario.objects.get(id=usuario)
            if User.objects.filter(username=usuario1).exclude(username=exclude_id).exists():
                id2 = Usuario.objects.get(id=usuario)
                form = EditarRepositorForm(instance=id2)
                error_msg = "Ya existe el nombre de usuario "+usuario1+", use otro diferente."
                return render(request, 'editar_repositor.html',{'form':form,'usuario':id2,'error_msg':error_msg,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
            else:
                editar_personal = Usuario.objects.get(id=usuario)
                editar_personal.legajo = legajo
                editar_personal.save()
                editar_usuario = User.objects.get(username=editar_personal)
                editar_usuario.first_name = nombres
                editar_usuario.last_name = apellidos
                editar_usuario.email = email
                editar_usuario.username = usuario1
                editar_usuario.save()
                return redirect('/listado_repositores')
        else:
            id2 = Usuario.objects.get(id=usuario)
            print(id2.usuario.id)
            user_repositor = User.objects.get(id=id2.usuario.id)
            form = EditarRepositorForm(instance=id2)
            if user_repositor.is_active:   
                habilitado = True
            else:
                habilitado = False
        comercios = Comercio.objects.all()
        return render(request, 'editar_repositor.html',{'comercios':comercios,'form':form,'usuario':id2,'habilitado':habilitado,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})

@login_required(None,'login','/login/')
def repositor(request,usuario):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        usuarios = User.objects.filter(is_active=True,is_superuser=False, is_staff=False)
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(usuarios.count()*100/limite_usuarios)
        marcaciones= Marcacione.objects.filter(usuario=usuario).order_by('-fecha','-hora')
        if 'inicio' in request.GET and 'fin' in request.GET:
            f_inicio = request.GET.get('inicio')
            f_fin = request.GET.get('fin')
            inicio = datetime.strptime(f_inicio, "%d/%m/%Y")
            fin = datetime.strptime(f_fin, "%d/%m/%Y")
            marcaciones= Marcacione.objects.filter(usuario=usuario,fecha__range=[inicio, fin]).order_by('-fecha','-hora')
            filtro_msg = "Filtro de fechas entre el "+str(f_inicio)+" y el "+str(f_fin)
            
            return render(request, 'repositor.html',{'marcaciones':marcaciones,'filtro_msg':filtro_msg,'usuario':usuario,'inicio':inicio,'fin':fin,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
            
            
        return render(request, 'repositor.html',{'marcaciones':marcaciones,'usuario':usuario,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})
	
@login_required(None,'login','/login/')
def listado_marcaciones(request):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        usuarios = User.objects.filter(is_active=True,is_superuser=False, is_staff=False)
        limite_usuarios = ultimo_contrato.limite_usuarios
        porcentaje = int(usuarios.count()*100/limite_usuarios)
        marcaciones= Marcacione.objects.all().order_by('-fecha','-hora')
        if 'inicio' in request.GET and 'fin' in request.GET:
            f_inicio = request.GET.get('inicio')
            f_fin = request.GET.get('fin')
            inicio = datetime.strptime(f_inicio, "%d/%m/%Y")
            fin = datetime.strptime(f_fin, "%d/%m/%Y")
            marcaciones= Marcacione.objects.filter(fecha__range=[inicio, fin]).order_by('-fecha','-hora')
            filtro_msg = "Filtro de fechas entre el "+str(f_inicio)+" y el "+str(f_fin)
            return render(request, 'marcaciones.html',{'marcaciones':marcaciones,'filtro_msg':filtro_msg,'inicio':inicio,'fin':fin,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
        return render(request, 'marcaciones.html',{'marcaciones':marcaciones,'usuarios':usuarios,'porcentaje':porcentaje,'limite_usuarios':limite_usuarios})
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})

@login_required(None,'login','/login/')
def inhabilitar_repositor(request,usuario):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        # cambia el estado de User a inactivo
        inhabilitar_usuario = User.objects.get(id=usuario)
        inhabilitar_usuario.is_active = False
        inhabilitar_usuario.save()
        return redirect('/listado_repositores')
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})

@login_required(None,'login','/login/')
def habilitar_repositor(request,usuario):
    ultimo_contrato = Contrato.objects.latest('id')
    hoy = date.today()
    user = request.user
    if ultimo_contrato.fecha_caducidad >= hoy or user.is_superuser:
        # cambia el estado de User a activo
        habilitar_usuario = User.objects.get(id=usuario)
        habilitar_usuario.is_active = True
        habilitar_usuario.save()
        return redirect('/listado_repositores')
    else:
        logout(request)
        state = "Tu cuenta ha caducado, por favor ponte en contacto con el administrador."
        print(state)
        return render(request,'login.html',{'state':state})

#Nuestra clase hereda de la vista genérica TemplateView
class ReportePersonasExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        if 'inicio' in request.GET and 'fin' in request.GET:
            try:
                print("Tiene Repositor y tiene Filtro")
                q = str(request.GET['repositor'])
                inicio = request.GET.get('inicio')
                fin = request.GET.get('fin')
                marcaciones = Marcacione.objects.filter(usuario=q,fecha__range=[inicio, fin]).order_by('-fecha','-hora')
                filtro_msg = "Filtro de fechas entre el "+str(inicio)+" y el "+str(fin)
            except:
                print("NO tiene Repositor y SI tiene Filtro")
                inicio = request.GET.get('inicio')
                fin = request.GET.get('fin')
                marcaciones = Marcacione.objects.filter(fecha__range=[inicio, fin]).order_by('-fecha','-hora')
                filtro_msg = "Filtro de fechas entre el "+str(inicio)+" y el "+str(fin)
        else:
            try:
                print("Tiene Repositor y NO tiene Filtro")
                q = str(request.GET['repositor'])
                marcaciones = Marcacione.objects.filter(usuario=q).order_by('-fecha','-hora')
            except:
                print("NO tene Repositor y NO tiene Filtro")
                marcaciones = Marcacione.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE REPOSITORES'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:J2')
		# Python types will automatically be converted
        import datetime
        ws['B3'] = datetime.datetime.now()
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B4'] = 'Legajo'
        ws['C4'] = 'Nombres'
        ws['D4'] = 'Apellidos'
        ws['E4'] = 'Fecha'
        ws['F4'] = 'Hora'
        ws['G4'] = 'Observaciones'
        ws['H4'] = 'Lugar'
        ws['I4'] = 'Zona'  
        ws['J4'] = 'Estado'       
        cont=5
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for marcacion in marcaciones:
            ws.cell(row=cont,column=2).value = marcacion.usuario.username
            ws.cell(row=cont,column=3).value = marcacion.usuario.first_name
            ws.cell(row=cont,column=4).value = marcacion.usuario.last_name
            ws.cell(row=cont,column=5).value = marcacion.fecha
            ws.cell(row=cont,column=6).value = marcacion.hora
            ws.cell(row=cont,column=7).value = marcacion.observaciones
            ws.cell(row=cont,column=8).value = marcacion.lugar
            ws.cell(row=cont,column=9).value = marcacion.zona
            if marcacion.estado == "0":
                ws.cell(row=cont,column=10).value = "Entrada"
            else:
                ws.cell(row=cont,column=10).value = "Salida"
            cont = cont + 1
        #img = Image('static/img/logo-letra.png')
        #ws.add_image(img, 'I3')
        ws.merge_cells('I3:J3')
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteRepositores.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
 
def mi_error_404(request):
    nombre_template = '404.html'
 
    return page_not_found(request, template_name=nombre_template)


	
